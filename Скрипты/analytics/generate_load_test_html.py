#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate HTML report from load testing Excel (Was/Became blocks by optimization).

Reads: Проекты/Результаты нагрузочного тестирования. Выводы +Исполнение - Замеры.xlsx
Writes: Проекты/load_test_results_report.html
"""

from pathlib import Path
import re
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = PROJECT_ROOT / "Проекты" / "Результаты нагрузочного тестирования. Выводы +Исполнение - Замеры.xlsx"
OUT_HTML = PROJECT_ROOT / "Проекты" / "load_test_results_report.html"

# System headers in column A
SYSTEM_NAMES = ("AvancoreDU", "AvancoreMO", "AvancoreFIN", "AvancoreFO")
TABLE_HEADER_ROW = "Наименование сценария"


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return (str(v).strip() if v is not None else "") or ""


def _is_note_row(name):
    """Skip footer/note rows that are not real scenarios."""
    if not name:
        return True
    n = name.strip()
    return n.startswith("Для ") or n.startswith("Владелец ") or n == "На старте"


def _normalize_time(s):
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_sheet(ws):
    """Parse Sheet1 into blocks: { system_name: [ { name, volume, target, result, recommendation, ticket }, ... ] }."""
    blocks = {}
    current_system = None
    current_scenario = None
    scenario_volumes = []

    for r in range(1, ws.max_row + 1):
        a1 = _cell(ws, r, 1)
        a2 = _cell(ws, r, 2)
        a3 = _cell(ws, r, 3)
        a4 = _cell(ws, r, 4)
        a5 = _cell(ws, r, 5)
        a6 = _cell(ws, r, 6)
        a7 = _cell(ws, r, 7)
        a8 = _cell(ws, r, 8)

        if a1 in SYSTEM_NAMES:
            current_system = a1
            blocks[current_system] = []
            current_scenario = None
            scenario_volumes = []
            continue
        if a1 == TABLE_HEADER_ROW:
            continue
        if not current_system:
            continue

        if a1:
            if current_scenario and current_scenario.get("name"):
                vol = " ; ".join(filter(None, scenario_volumes))
                if vol:
                    current_scenario["volume"] = vol
                name = current_scenario["name"]
                if not _is_note_row(name):
                    blocks[current_system].append(current_scenario)
            current_scenario = {
                "name": a1,
                "volume": a2,
                "target": _normalize_time(a3),
                "result": _normalize_time(a4),
                "recommendation": (a5 or "").strip(),
                "extra": (a6 or "").strip(),
                "ticket": (a8 or "").strip(),
            }
            scenario_volumes = [a2] if a2 else []
        else:
            if current_scenario is not None:
                if a2:
                    scenario_volumes.append(a2)
                if a3 and not current_scenario.get("target"):
                    current_scenario["target"] = _normalize_time(a3)
                if a4 and not current_scenario.get("result"):
                    current_scenario["result"] = _normalize_time(a4)
                if a5:
                    current_scenario["recommendation"] = (current_scenario.get("recommendation") or "") + " " + a5
                if a6:
                    current_scenario["extra"] = (current_scenario.get("extra") or "") + " " + a6
                if a8:
                    current_scenario["ticket"] = (current_scenario.get("ticket") or "") + " " + a8

    if current_scenario and current_scenario.get("name"):
        vol = " ; ".join(filter(None, scenario_volumes))
        if vol:
            current_scenario["volume"] = vol
        if not _is_note_row(current_scenario["name"]):
            blocks[current_system].append(current_scenario)

    return blocks


def status_class(scenario):
    """Return status: ok / warn / fail / unknown."""
    res = (scenario.get("result") or "").lower()
    tgt = scenario.get("target") or ""
    if "не завершился" in res or "не завершился" in (scenario.get("result") or ""):
        return "fail"
    if not res and not tgt:
        return "unknown"
    if not tgt:
        return "info"
    if not res:
        return "unknown"
    # Simple heuristic: if result contains "мин" and target "час" - often worse
    if "прогноз" in res:
        return "warn"
    return "info"


def escape_html(s):
    if not s:
        return ""
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_card(scenario, sys_name):
    status = status_class(scenario)
    name = escape_html(scenario.get("name") or "")
    volume = escape_html(scenario.get("volume") or "")
    target = escape_html(scenario.get("target") or "—")
    result = escape_html(scenario.get("result") or "—")
    rec = escape_html(scenario.get("recommendation") or "").strip()
    extra = escape_html(scenario.get("extra") or "").strip()
    ticket = escape_html(scenario.get("ticket") or "").strip()

    status_label = {"ok": "Уложились", "warn": "Превышение", "fail": "Не завершён", "info": "Результат"}.get(
        status, "Результат"
    )

    parts = [
        '<article class="card card--%s">',
        '  <header class="card__header">',
        '    <h3 class="card__title">%s</h3>',
        '    <span class="card__badge card__badge--%s">%s</span>',
        "  </header>",
        '  <div class="card__body">',
        '    <div class="block block--was-stalo">',
        '      <div class="block__col block__col--was">',
        '        <div class="block__label">Было (цель)</div>',
        "        <div class=\"block__value block__value--target\">Объём: %s</div>",
        "        <div class=\"block__value block__value--target\">Целевое время: %s</div>",
        "      </div>",
        '      <div class="block__col block__col--stalo">',
        '        <div class="block__label">Стало (результат)</div>',
        "        <div class=\"block__value block__value--result\">%s</div>",
        "      </div>",
        "    </div>",
    ]

    if rec or extra or ticket:
        parts.extend(
            [
                '    <div class="card__meta">',
            ]
        )
        if rec:
            parts.append('      <p class="card__rec"><strong>Рекомендация:</strong> %s</p>' % rec)
        if extra:
            parts.append('      <p class="card__extra">%s</p>' % extra)
        if ticket:
            parts.append('      <p class="card__ticket"><strong>Задача:</strong> %s</p>' % ticket)
        parts.append("    </div>")

    parts.extend(["  </div>", "</article>"])

    return "\n".join(parts) % (status, name, status, status_label, volume, target, result)


def build_html(blocks):
    system_titles = {
        "AvancoreDU": "Avancore DU (управление активами)",
        "AvancoreMO": "Avancore MO (Middle Office)",
        "AvancoreFIN": "Avancore FIN (финансы)",
        "AvancoreFO": "Avancore FO (Front Office)",
    }

    toc_parts = []
    cards_html = []
    for sys_name in SYSTEM_NAMES:
        scenarios = blocks.get(sys_name, [])
        if not scenarios:
            continue
        title = system_titles.get(sys_name, sys_name)
        toc_parts.append(
            '    <a class="toc__link" href="#%s">%s <span class="toc__count">%d</span></a>'
            % (escape_html(sys_name), escape_html(title), len(scenarios))
        )
        cards_html.append('<section class="section" id="%s">' % escape_html(sys_name))
        cards_html.append('  <h2 class="section__title">%s</h2>' % escape_html(title))
        cards_html.append('  <div class="cards">')
        for sc in scenarios:
            cards_html.append(render_card(sc, sys_name))
        cards_html.append("  </div>")
        cards_html.append("</section>")

    nav_html = ""
    if toc_parts:
        nav_html = (
            '\n  <nav class="toc" aria-label="Содержание">\n'
            '    <h2 class="toc__title">Блоки оптимизации</h2>\n'
            + "\n".join(toc_parts)
            + "\n  </nav>\n"
        )

    html = """<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Результаты нагрузочного тестирования. Было / Стало</title>
  <style>
    :root {
      --color-bg: #0f1419;
      --color-surface: #1a2332;
      --color-surface2: #243044;
      --color-text: #e6edf3;
      --color-text-muted: #8b9cb3;
      --color-success: #28a745;
      --color-danger: #dc3545;
      --color-warning: #ffc107;
      --color-info: #17a2b8;
      --font-sans: 'Segoe UI', system-ui, -apple-system, sans-serif;
      --radius: 12px;
      --shadow: 0 4px 24px rgba(0,0,0,0.25);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      padding: 2rem 1rem 4rem;
      font-family: var(--font-sans);
      font-size: 16px;
      line-height: 1.5;
      color: var(--color-text);
      background: var(--color-bg);
    }
    .page-header {
      text-align: center;
      margin-bottom: 2rem;
      padding-bottom: 2rem;
      border-bottom: 1px solid var(--color-surface2);
    }
    .page-header h1 {
      font-size: 1.75rem;
      font-weight: 700;
      margin: 0 0 0.5rem;
      letter-spacing: -0.02em;
    }
    .page-header p {
      color: var(--color-text-muted);
      margin: 0;
      font-size: 0.95rem;
    }
    .toc {
      max-width: 960px;
      margin: 0 auto 2.5rem;
      padding: 1.25rem 1.5rem;
      background: var(--color-surface);
      border-radius: var(--radius);
      border: 1px solid var(--color-surface2);
    }
    .toc__title {
      font-size: 0.85rem;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      color: var(--color-text-muted);
      margin: 0 0 1rem;
    }
    .toc__link {
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 0.5rem 0;
      color: var(--color-info);
      text-decoration: none;
      border-bottom: 1px solid var(--color-surface2);
      font-weight: 500;
    }
    .toc__link:hover { color: var(--color-text); }
    .toc__link:last-child { border-bottom: none; }
    .toc__count {
      font-size: 0.8rem;
      color: var(--color-text-muted);
      font-weight: 400;
    }
    .section {
      max-width: 960px;
      margin: 0 auto 3rem;
    }
    .section__title {
      font-size: 1.35rem;
      font-weight: 600;
      margin: 0 0 1.5rem;
      color: var(--color-info);
    }
    .cards {
      display: flex;
      flex-direction: column;
      gap: 1.25rem;
    }
    .card {
      background: var(--color-surface);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow: hidden;
      border: 1px solid var(--color-surface2);
      border-left-width: 4px;
      border-left-style: solid;
    }
    .card--ok { border-left-color: var(--color-success); }
    .card--warn { border-left-color: var(--color-warning); }
    .card--fail { border-left-color: var(--color-danger); }
    .card--info, .card--unknown { border-left-color: var(--color-info); }
    .card__header {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 1rem;
      padding: 1rem 1.25rem;
      background: var(--color-surface2);
      border-bottom: 1px solid rgba(255,255,255,0.06);
    }
    .card__title {
      margin: 0;
      font-size: 1.05rem;
      font-weight: 600;
      flex: 1;
    }
    .card__badge {
      flex-shrink: 0;
      padding: 0.25rem 0.6rem;
      border-radius: 999px;
      font-size: 0.75rem;
      font-weight: 600;
      text-transform: uppercase;
    }
    .card__badge--ok { background: var(--color-success); color: #fff; }
    .card__badge--warn { background: var(--color-warning); color: #1a2332; }
    .card__badge--fail { background: var(--color-danger); color: #fff; }
    .card__badge--info { background: var(--color-info); color: #fff; }
    .card__badge--unknown { background: var(--color-text-muted); color: #fff; }
    .card__body { padding: 1.25rem; }
    .block--was-stalo {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 1.5rem;
      margin-bottom: 1rem;
    }
    @media (max-width: 640px) {
      .block--was-stalo { grid-template-columns: 1fr; }
    }
    .block__label {
      font-size: 0.7rem;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      color: var(--color-text-muted);
      margin-bottom: 0.35rem;
    }
    .block__value {
      font-size: 1rem;
      font-weight: 500;
    }
    .block__col--was .block__value--target { color: var(--color-text-muted); }
    .block__col--stalo .block__value--result { color: var(--color-text); }
    .card__meta { margin-top: 1rem; padding-top: 1rem; border-top: 1px solid var(--color-surface2); }
    .card__meta p { margin: 0 0 0.5rem; font-size: 0.9rem; color: var(--color-text-muted); }
    .card__meta p:last-child { margin-bottom: 0; }
    .card__rec strong, .card__ticket strong { color: var(--color-text); }
    @media print {
      body { background: #fff; color: #111; }
      .toc { background: #f5f5f5; border-color: #ddd; }
      .toc__link { color: #06c; }
      .section__title { color: #06c; }
      .card { background: #fff; border: 1px solid #ddd; box-shadow: none; }
      .card__header { background: #f5f5f5; }
      .block__col--was .block__value--target { color: #444; }
      .card__meta p { color: #444; }
    }
  </style>
</head>
<body>
  <header class="page-header">
    <h1>Результаты нагрузочного тестирования</h1>
    <p>Замеры по блокам оптимизации: Было (цель) / Стало (результат)</p>
  </header>
"""
    html += nav_html
    html += "\n".join(cards_html)
    html += """
</body>
</html>
"""
    return html


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    blocks = parse_sheet(ws)
    wb.close()

    html = build_html(blocks)
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    print("Written: " + str(OUT_HTML))


if __name__ == "__main__":
    main()
